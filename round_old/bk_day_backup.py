#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import akshare as ak
import pandas as pd
import os
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Alignment, Font
import html
import re

# ================== 配置 ==================
class Config:
    TOP_K = 20

    OUTPUT_XLSX = "html/bk_day_top.xlsx"
    HTML_TOP = "html/bk_day_info.html"
    HTML_ALL = "html/bk_day_all.html"
    HTML_RANK = "html/bk_day_rank.html"

    GEN_HTML_TOPBOTTOM = True
    GEN_HTML_ALLDATA = True
    GEN_HTML_ALLDATA_RANK = True

    # Excel / HTML 样式
    ROW_HEIGHT = 17
    FONT_SIZE = 10
    HTML_INFO_FONT_SIZE = 8
    RANK_COL_WIDTH = 20
    DATE_COL_WIDTH = 55
    INFO_COL_WIDTH = 30
    REMARK_COL_WIDTH = 30

    SHOW_INFO_COL = True
    SHOW_REMARK_COL = False

# ================== 工具函数 ==================
def pct_to_level(pct):
    level = min(int(abs(pct) / 0.5) + 1, 10)
    return level

def make_fill(pct):
    red = {1:"FFECEC",2:"FFD9D9",3:"FFC6C6",4:"FFB3B3",5:"FF9999",
           6:"FF8080",7:"FF6666",8:"FF4D4D",9:"FF3333",10:"FF1A1A"}
    green = {1:"E9F7EF",2:"D3F0DE",3:"BCE9CE",4:"A6E2BD",5:"8FDAB3",
             6:"78D2A9",7:"61C99F",8:"4ABF95",9:"33B58B",10:"1CAB80"}
    level = pct_to_level(pct)
    color = red[level] if pct >=0 else green[level]
    return PatternFill(fill_type="solid", fgColor=color)

def safe_int(v):
    return 0 if pd.isna(v) else int(v)

def shorten_col(col):
    if re.match(r"\d{4}-\d{2}-\d{2}$", col):
        return col[5:]
    if col.endswith("_info"):
        return "info"
    if col.endswith("_remark"):
        return "remark"
    return col

def get_latest_trade_date():
    today = datetime.now().strftime("%Y-%m-%d")
    try:
        trade_cal = ak.tool_trade_date_hist_sina()
        trade_dates = trade_cal['trade_date'].dropna().tolist()
        trade_dates = [str(d) if not hasattr(d,"strftime") else d.strftime("%Y-%m-%d") for d in trade_dates]
        if today in trade_dates:
            return today
        for d in reversed(trade_dates):
            if d < today:
                return d
    except Exception as e:
        print(f"[WARN] 获取交易日失败，使用今天: {e}")
    return today

def format_html_info(val):
    if not isinstance(val, str) or not val.strip():
        return "<td></td>"
    parts = val.split("/")
    display_val = parts[0].strip() if parts else val.strip()
    try:
        pct = float(display_val)
        color_class = "pos" if pct >=0 else "neg"
    except:
        color_class = ""
    return f'<td class="{color_class}">{html.escape(display_val)}</td>'

# ================== 数据处理 ==================
def fetch_board_data(top_k):
    df = ak.stock_board_industry_name_em()
    df = df.rename(columns={
        "板块名称": "board_name",
        "涨跌幅": "pct",
        "换手率": "turnover",
        "上涨家数": "up_count",
        "下跌家数": "down_count",
        "领涨股票": "leader_stock"
    })
    df["pct"] = pd.to_numeric(df["pct"], errors="coerce")
    df = df.dropna(subset=["pct"])
    df_sorted = df.sort_values("pct", ascending=False).reset_index(drop=True)

    top = df_sorted.head(top_k)
    bottom = df_sorted.tail(top_k)
    empty = pd.DataFrame([{c:"" for c in df_sorted.columns}])
    merged = pd.concat([top, empty, bottom], ignore_index=True)

    name_col, info_col, rank_col, pct_map = [], [], [], {}
    total_len = len(df_sorted)

    for idx, row in merged.iterrows():
        if row.isnull().all() or all(row==""):
            name_col.append("")
            info_col.append("")
            rank_col.append("")
            pct_map[idx+1] = None
        else:
            name_col.append(row["board_name"])
            info = f"{row['pct']:.2f} / {row.get('turnover',0):.2f} / {safe_int(row.get('up_count',0))} / {safe_int(row.get('down_count',0))} / {row.get('leader_stock','')}"
            info_col.append(info)
            if idx < top_k:
                rank_col.append(idx+1)
            else:
                rank_col.append(total_len - (top_k - (idx-top_k-1)) + 1)
            pct_map[idx+1] = row["pct"]

    return merged, name_col, info_col, rank_col, pct_map, df_sorted

# ================== Excel 操作 ==================
def update_excel(today, name_col, info_col, rank_col, pct_map, df_all_sorted, config:Config):
    max_len = len(name_col)
    # 新建 Excel
    if not os.path.exists(config.OUTPUT_XLSX):
        wb = Workbook()
        ws = wb.active
        ws.title = "TopBottom"
        ws.append(["rank", today, f"{today}_info", f"{today}_remark"])
        for i in range(max_len):
            ws.append([rank_col[i], name_col[i], info_col[i], ""])
        for cell in ws[1]:
            cell.font = Font(bold=True)
        wb.save(config.OUTPUT_XLSX)

    wb = load_workbook(config.OUTPUT_XLSX)

    def update_sheet(ws, names, infos, ranks, pct_map, max_len, today):
        header = [c.value for c in ws[1]]
        if today in header:
            idx = header.index(today)
            ws.delete_cols(idx+1, 3)
        ws.insert_cols(2,3)
        ws.cell(row=1, column=2, value=today)
        ws.cell(row=1, column=3, value=f"{today}_info")
        ws.cell(row=1, column=4, value=f"{today}_remark")

        for i in range(max_len):
            ws.cell(row=i+2, column=1, value=ranks[i])
            ws.cell(row=i+2, column=2, value=names[i])
            ws.cell(row=i+2, column=3, value=infos[i])
            ws.cell(row=i+2, column=4, value="")

        ws.freeze_panes = "B2"
        for i in range(2,2+max_len):
            ws.cell(row=i,column=1).alignment = Alignment(horizontal="center")
            pct = pct_map.get(i-1)
            if pct is not None:
                ws.cell(row=i,column=3).fill = make_fill(pct)

        for col_cells in ws.columns:
            col_letter = col_cells[0].column_letter
            max_len_col = max(len(str(c.value)) if c.value else 0 for c in col_cells)
            ws.column_dimensions[col_letter].width = max(6, max_len_col+2)

    # 更新 TopBottom
    ws_tb = wb["TopBottom"] if "TopBottom" in wb.sheetnames else wb.create_sheet("TopBottom")
    update_sheet(ws_tb, name_col, info_col, rank_col, pct_map, max_len, today)

    # 更新 AllData
    name_all = df_all_sorted["board_name"].tolist()
    info_all = [
        f"{row['pct']:.2f} / {row.get('turnover',0):.2f} / {safe_int(row.get('up_count',0))} / {safe_int(row.get('down_count',0))} / {row.get('leader_stock','')}"
        for _, row in df_all_sorted.iterrows()
    ]
    rank_all = list(range(1,len(df_all_sorted)+1))
    pct_all_map = {i+1: df_all_sorted.iloc[i]["pct"] for i in range(len(df_all_sorted))}
    ws_all = wb["AllData"] if "AllData" in wb.sheetnames else wb.create_sheet("AllData")
    update_sheet(ws_all, name_all, info_all, rank_all, pct_all_map, len(df_all_sorted), today)

    wb.save(config.OUTPUT_XLSX)
    print(f"[OK] Excel 已更新：{config.OUTPUT_XLSX}")
    return wb

# ================== HTML 生成 ==================
STYLE = f"""
<style>
body {{ font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",Arial,sans-serif; background:#f7f7f7; margin:0; padding:0; }}
.table-wrap {{ overflow-x:auto; background:#fff; padding:0; border-radius:6px; display:inline-block; }}
table {{ border-collapse:collapse; white-space:nowrap; font-size:{Config.FONT_SIZE}px; }}
th,td {{ border:1px solid #ddd; padding:1px 2px; vertical-align:middle; height:{Config.ROW_HEIGHT}px; }}
th {{ background:#333; color:#fff; position:sticky; top:0; z-index:2; }}
th[colclass="rank"], td[colclass="rank"] {{ width:{Config.RANK_COL_WIDTH}px; }}
th[colclass="date"], td[colclass="date"] {{ width:{Config.DATE_COL_WIDTH}px; }}
th[colclass="info"], td[colclass="info"] {{ width:{Config.INFO_COL_WIDTH}px; font-size:{Config.HTML_INFO_FONT_SIZE}px; }}
th[colclass="remark"], td[colclass="remark"] {{ width:{Config.REMARK_COL_WIDTH}px; }}
td.rank {{ text-align:center; font-weight:bold; background:#fafafa; }}
td.pos {{ color:#c00000; }}
td.neg {{ color:#006100; }}
td.remark {{ color:#555; font-size:12px; }}
td.bk {{ cursor:pointer; font-weight:bold; }}
td.bk.active-click {{ background:#ff9999 !important; }}
td.bk.active-hover {{ background:#ffcccc !important; }}
td.monday-col {{ background:#f0f0f0 !important; }}
</style>
"""

SCRIPT = """
<script>
let activeBk = null;
function toggleBk(el) {
    const name = el.getAttribute("data-bk");
    document.querySelectorAll("td.bk").forEach(td => td.classList.remove("active-click"));
    if (activeBk === name) { activeBk=null; return; }
    document.querySelectorAll(`td.bk[data-bk="${name}"]`).forEach(td => td.classList.add("active-click"));
    activeBk=name;
}
document.querySelectorAll("td.bk").forEach(td=>{
    td.addEventListener("mouseenter",()=>{const name=td.getAttribute("data-bk");document.querySelectorAll(`td.bk[data-bk="${name}"]`).forEach(el=>{if(!el.classList.contains("active-click"))el.classList.add("active-hover");});});
    td.addEventListener("mouseleave",()=>{const name=td.getAttribute("data-bk");document.querySelectorAll(`td.bk[data-bk="${name}"]`).forEach(el=>el.classList.remove("active-hover"));});
});
(function markMondayColumns() {
    const ths = document.querySelectorAll("th[colclass='date']");
    ths.forEach(th => {
        const colname = th.getAttribute("colname");
        if (!colname) return;
        const d = new Date(colname);
        if (isNaN(d.getTime())) return;
        if (d.getDay() === 1) {
            th.classList.add("monday-col");
            document.querySelectorAll(`td[colname='${colname}']`).forEach(td => td.classList.add("monday-col"));
        }
    });
})();
</script>
"""

def generate_html(df_excel, out_file, show_info=True, show_remark=False):
    filtered_cols = [c for c in df_excel.columns 
                     if not (c.endswith("_info") and not show_info) 
                     and not (c.endswith("_remark") and not show_remark)]
    display_cols = [shorten_col(c) for c in filtered_cols]

    # 表头
    thead = "<tr>"
    for col, display_col in zip(filtered_cols, display_cols):
        if col=="rank": colclass="rank"
        elif re.match(r"\d{4}-\d{2}-\d{2}$", col): colclass="date"
        elif col.endswith("_info"): colclass="info"
        elif col.endswith("_remark"): colclass="remark"
        else: colclass=""
        th_attr=f' colname="{col}" colclass="{colclass}"'
        thead += f"<th{th_attr}>{html.escape(display_col)}</th>"
    thead += "</tr>"

    # 表体
    tbody = ""
    for _, row in df_excel.iterrows():
        tbody += "<tr>"
        for col in filtered_cols:
            val = row[col]
            if col=="rank":
                tbody += f'<td class="rank">{int(val) if val!="" else ""}</td>'
                continue
            colclass="remark" if col.endswith("_remark") else "info" if col.endswith("_info") else "date" if re.match(r"\d{4}-\d{2}-\d{2}$",col) else ""
            td_attr=f' colname="{col}" colclass="{colclass}"'

            if colclass=="info":
                tbody += format_html_info(val)
                continue
            if colclass=="remark": 
                tbody += f'<td class="remark"{td_attr}>{html.escape(str(val))}</td>'
                continue
            if isinstance(val,str) and " / " not in val and not col.endswith("_info"):
                bk = html.escape(val.strip())
                if bk: tbody += f'<td class="bk"{td_attr} data-bk="{bk}" onclick="toggleBk(this)">{bk}</td>'
                else: tbody += f"<td{td_attr}></td>"
                continue
            tbody += f"<td{td_attr}>{html.escape(str(val))}</td>"
        tbody += "</tr>"

    table_html = f'<div class="table-wrap"><table><thead>{thead}</thead><tbody>{tbody}</tbody></table></div>'
    html_page = f"<!DOCTYPE html><html lang='zh-CN'><head><meta charset='utf-8'><title></title>{STYLE}</head><body>{table_html}{SCRIPT}</body></html>"
    Path(out_file).write_text(html_page, encoding="utf-8")
    print(f"[OK] HTML 已生成：{out_file}")

# ================== 主流程 ==================
def main():
    config = Config()
    today = get_latest_trade_date()
    print(f"[INFO] 使用交易日 {today} 数据更新...")

    merged, name_col, info_col, rank_col, pct_map, df_sorted = fetch_board_data(config.TOP_K)
    wb = update_excel(today, name_col, info_col, rank_col, pct_map, df_sorted, config)

    if config.GEN_HTML_TOPBOTTOM:
        df_excel = pd.read_excel(config.OUTPUT_XLSX, sheet_name="TopBottom").fillna("")
        generate_html(df_excel, config.HTML_TOP, show_info=config.SHOW_INFO_COL, show_remark=config.SHOW_REMARK_COL)

    if config.GEN_HTML_ALLDATA:
        df_all = pd.read_excel(config.OUTPUT_XLSX, sheet_name="AllData").fillna("")
        generate_html(df_all, config.HTML_ALL, show_info=True, show_remark=config.SHOW_REMARK_COL)

    if config.GEN_HTML_ALLDATA_RANK:
        df_rank = pd.read_excel(config.OUTPUT_XLSX, sheet_name="AllData").fillna("")
        # 仅保留 rank / 名称 / 日期列 / info
        filtered_cols = [c for c in df_rank.columns if c=="rank" or re.match(r"\d{4}-\d{2}-\d{2}$",c) or c.endswith("_info")]
        generate_html(df_rank[filtered_cols], config.HTML_RANK, show_info=True)

if __name__=="__main__":
    main()
