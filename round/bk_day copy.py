#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import akshare as ak
import pandas as pd
import os
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Font
import html
import re

# ================== 参数 ==================
TOP_K = 20
OUTPUT_XLSX = "html/bk_day.xlsx"
OUT_HTML = "html/bk_day_top.html"

# ================== HTML 输出控制 ==================
GEN_HTML_TOPBOTTOM = True      # 生成 bk_day.html（现有）
GEN_HTML_ALLDATA = True        # 生成 bk_day_all.html（现有）
GEN_HTML_ALLDATA_RANK = True   # 新增：基于 AllData 的纯排名 html


ROW_HEIGHT = 17
FONT_SIZE = 10
HTML_INFO_FONT_SIZE = 8
RANK_COL_WIDTH = 20
DATE_COL_WIDTH = 55
INFO_COL_WIDTH = 30
REMARK_COL_WIDTH = 30

SHOW_INFO_COL = True
SHOW_REMARK_COL = False

# ================== 工具函数 ==================
def pct_to_level(pct):
    a = abs(pct)
    level = int(a / 0.5) + 1
    if level > 10:
        level = 10
    return level

def make_fill(pct):
    level = pct_to_level(pct)
    red = {1:"FFECEC",2:"FFD9D9",3:"FFC6C6",4:"FFB3B3",5:"FF9999",
           6:"FF8080",7:"FF6666",8:"FF4D4D",9:"FF3333",10:"FF1A1A"}
    green = {1:"E9F7EF",2:"D3F0DE",3:"BCE9CE",4:"A6E2BD",5:"8FDAB3",
             6:"78D2A9",7:"61C99F",8:"4ABF95",9:"33B58B",10:"1CAB80"}
    color = red[level] if pct >=0 else green[level]
    return PatternFill(fill_type="solid", fgColor=color)

def safe_int(v):
    if pd.isna(v):
        return 0
    return int(v)

def shorten_col(col):
    if re.match(r"\d{4}-\d{2}-\d{2}$", col):
        return col[5:]
    if col.endswith("_info"):
        return "info"
    if col.endswith("_remark"):
        return "remark"
    return col

def get_latest_trade_date():
    today = datetime.now().strftime("%Y-%m-%d")
    try:
        trade_cal = ak.tool_trade_date_hist_sina()
        trade_dates = trade_cal[trade_cal['trade_date'].notna()]['trade_date'].tolist()
        trade_dates = [d.strftime("%Y-%m-%d") if hasattr(d, "strftime") else str(d) for d in trade_dates]
        if today in trade_dates:
            return today
        for d in reversed(trade_dates):
            if d < today:
                return d
    except Exception as e:
        print(f"[WARN] 获取交易日失败，使用今天: {e}")
    return today

# HTML info 列显示：仅涨跌幅
def format_html_info(val):
    if not isinstance(val, str) or not val.strip():
        return "<td></td>"
    parts = val.split("/")
    display_val = parts[0].strip() if parts else val.strip()
    try:
        pct = float(display_val)
    except:
        pct = 0
    color_class = "pos" if pct >= 0 else "neg"
    return f'<td class="{color_class}">{html.escape(display_val)}</td>'

# ================== 主流程 ==================
def main():
    today = get_latest_trade_date()
    print(f"[INFO] 使用交易日 {today} 数据更新...")

    # ---------- 获取数据 ----------
    df = ak.stock_board_industry_name_em()
    df = df.rename(columns={
        "板块名称": "board_name",
        "涨跌幅": "pct",
        "换手率": "turnover",
        "上涨家数": "up_count",
        "下跌家数": "down_count",
        "领涨股票": "leader_stock"
    })
    df["pct"] = pd.to_numeric(df["pct"], errors="coerce")
    df = df.dropna(subset=["pct"])
    
    df_sorted = df.sort_values("pct", ascending=False).reset_index(drop=True)
    top = df_sorted.head(TOP_K)
    bottom = df_sorted.tail(TOP_K)
    empty = pd.DataFrame([{c:"" for c in df_sorted.columns}])
    merged = pd.concat([top, empty, bottom], ignore_index=True)
    
    name_col, info_col, rank_col, pct_map = [], [], [], {}
    total_len = len(df_sorted)
    for idx, row in merged.iterrows():
        if row.isnull().all() or all(row==""):
            name_col.append("")
            info_col.append("")
            rank_col.append("")
            pct_map[idx+1] = None
        else:
            name_col.append(row["board_name"])
            info = f"{row['pct']:.2f} / {row.get('turnover',0):.2f} / {safe_int(row.get('up_count',0))} / {safe_int(row.get('down_count',0))} / {row.get('leader_stock','')}"
            info_col.append(info)
            if idx < TOP_K:
                rank_col.append(idx+1)
            else:
                rank_col.append(total_len - (TOP_K - (idx-TOP_K-1)) + 1)
            pct_map[idx+1] = row["pct"]

    max_len = len(name_col)

    # ---------- Excel 更新 ----------
    from openpyxl import Workbook
    if not os.path.exists(OUTPUT_XLSX):
        wb = Workbook()
        ws = wb.active
        ws.title = "TopBottom"
        ws.append(["rank", today, f"{today}_info", f"{today}_remark"])
        for i in range(max_len):
            ws.append([rank_col[i], name_col[i], info_col[i], ""])
        for cell in ws[1]:
            cell.font = Font(bold=True)
        wb.save(OUTPUT_XLSX)

    wb = load_workbook(OUTPUT_XLSX)

    # ---------- TopBottom 页 ----------
    ws = wb["TopBottom"] if "TopBottom" in wb.sheetnames else wb.create_sheet("TopBottom")
    header = [c.value for c in ws[1]]
    if today in header:
        idx = header.index(today)
        ws.delete_cols(idx+1, 3)
    ws.insert_cols(2, 3)
    ws.cell(row=1, column=2, value=today)
    ws.cell(row=1, column=3, value=f"{today}_info")
    ws.cell(row=1, column=4, value=f"{today}_remark")
    for i in range(max_len):
        ws.cell(row=i+2, column=1, value=rank_col[i])
        ws.cell(row=i+2, column=2, value=name_col[i])
        ws.cell(row=i+2, column=3, value=info_col[i])
        ws.cell(row=i+2, column=4, value="")
    ws.freeze_panes = "B2"
    for i in range(2, 2+max_len):
        ws.cell(row=i, column=1).alignment = Alignment(horizontal="center")
        pct = pct_map.get(i-1)
        if pct is not None:
            ws.cell(row=i, column=3).fill = make_fill(pct)
    for col_cells in ws.columns:
        col_letter = col_cells[0].column_letter
        max_len_col = max(len(str(c.value)) if c.value else 0 for c in col_cells)
        ws.column_dimensions[col_letter].width = max(6, max_len_col+2)

    # ---------- AllData 页 ----------
    name_all = df_sorted["board_name"].tolist()
    info_all = [
        f"{row['pct']:.2f} / {row.get('turnover',0):.2f} / {safe_int(row.get('up_count',0))} / {safe_int(row.get('down_count',0))} / {row.get('leader_stock','')}"
        for _, row in df_sorted.iterrows()
    ]
    rank_all = list(range(1, total_len+1))
    pct_all_map = {i+1: df_sorted.iloc[i]["pct"] for i in range(total_len)}
    ws_all = wb["AllData"] if "AllData" in wb.sheetnames else wb.create_sheet("AllData")
    header = [c.value for c in ws_all[1]] if ws_all.max_row>0 else []
    if today in header:
        idx = header.index(today)
        ws_all.delete_cols(idx+1, 3)
    ws_all.insert_cols(2,3)
    ws_all.cell(row=1,column=2,value=today)
    ws_all.cell(row=1,column=3,value=f"{today}_info")
    ws_all.cell(row=1,column=4,value=f"{today}_remark")
    for i in range(total_len):
        row_idx = i+2
        if row_idx>ws_all.max_row:
            ws_all.append([rank_all[i], name_all[i], info_all[i], ""])
        else:
            ws_all.cell(row=row_idx,column=1,value=rank_all[i])
            ws_all.cell(row=row_idx,column=2,value=name_all[i])
            ws_all.cell(row=row_idx,column=3,value=info_all[i])
            ws_all.cell(row=row_idx,column=4,value="")
    ws_all.freeze_panes = "B2"
    for i in range(2,2+total_len):
        ws_all.cell(row=i,column=1).alignment = Alignment(horizontal="center")
        pct = pct_all_map.get(i-1)
        if pct is not None:
            ws_all.cell(row=i,column=3).fill = make_fill(pct)
    for col_cells in ws_all.columns:
        col_letter = col_cells[0].column_letter
        max_len_col = max(len(str(c.value)) if c.value else 0 for c in col_cells)
        ws_all.column_dimensions[col_letter].width = max(6, max_len_col+2)

    wb.save(OUTPUT_XLSX)
    print(f"[OK] Excel 已更新：{OUTPUT_XLSX}")

    # ---------- HTML 生成 (bk_day.html info只显示涨跌幅) ----------
    df_excel = pd.read_excel(OUTPUT_XLSX, sheet_name="TopBottom").fillna("")
    filtered_cols = []
    for col in df_excel.columns:
        if col.endswith("_info") and not SHOW_INFO_COL: continue
        if col.endswith("_remark") and not SHOW_REMARK_COL: continue
        filtered_cols.append(col)
    display_cols = [shorten_col(c) for c in filtered_cols]

    STYLE = f"""
    <style>
    body {{ font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",Arial,sans-serif; background:#f7f7f7; margin:0; padding:0; }}
    .table-wrap {{ overflow-x:auto; background:#fff; padding:0; border-radius:6px; display:inline-block; }}
    table {{ border-collapse:collapse; white-space:nowrap; font-size:{FONT_SIZE}px; }}
    th,td {{ border:1px solid #ddd; padding:1px 2px; vertical-align:middle; height:{ROW_HEIGHT}px; }}
    th {{ background:#333; color:#fff; position:sticky; top:0; z-index:2; }}
    th[colclass="rank"], td[colclass="rank"] {{ width:{RANK_COL_WIDTH}px; }}
    th[colclass="date"], td[colclass="date"] {{ width:{DATE_COL_WIDTH}px; }}
    th[colclass="info"], td[colclass="info"] {{ width:{INFO_COL_WIDTH}px; font-size:{HTML_INFO_FONT_SIZE}px; }}
    th[colclass="remark"], td[colclass="remark"] {{ width:{REMARK_COL_WIDTH}px; }}
    td.rank {{ text-align:center; font-weight:bold; background:#fafafa; }}
    td.pos {{ color:#c00000; }}
    td.neg {{ color:#006100; }}
    td.remark {{ color:#555; font-size:12px; }}
    td.bk {{ cursor:pointer; font-weight:bold; }}
    td.bk.active-click {{ background:#ff9999 !important; }}
    td.bk.active-hover {{ background:#ffcccc !important; }}
    td.monday-col {{ background:#f0f0f0 !important; }}
    </style>
    """

    SCRIPT = """
    <script>
    let activeBk = null;
    function toggleBk(el) {
        const name = el.getAttribute("data-bk");
        document.querySelectorAll("td.bk").forEach(td => td.classList.remove("active-click"));
        if (activeBk === name) { activeBk=null; return; }
        document.querySelectorAll(`td.bk[data-bk="${name}"]`).forEach(td => td.classList.add("active-click"));
        activeBk=name;
    }
    document.querySelectorAll("td.bk").forEach(td=>{
        td.addEventListener("mouseenter",()=>{const name=td.getAttribute("data-bk");document.querySelectorAll(`td.bk[data-bk="${name}"]`).forEach(el=>{if(!el.classList.contains("active-click"))el.classList.add("active-hover");});});
        td.addEventListener("mouseleave",()=>{const name=td.getAttribute("data-bk");document.querySelectorAll(`td.bk[data-bk="${name}"]`).forEach(el=>el.classList.remove("active-hover"));});
    });
    (function markMondayColumns() {
        const ths = document.querySelectorAll("th[colclass='date']");
        ths.forEach(th => {
            const colname = th.getAttribute("colname");
            if (!colname) return;
            const d = new Date(colname);
            if (isNaN(d.getTime())) return;
            if (d.getDay() === 1) {
                th.classList.add("monday-col");
                document.querySelectorAll(`td[colname='${colname}']`).forEach(td => td.classList.add("monday-col"));
            }
        });
    })();
    </script>
    """

    # 表头
    thead = "<tr>"
    for col, display_col in zip(filtered_cols, display_cols):
        if col=="rank": colclass="rank"
        elif re.match(r"\d{4}-\d{2}-\d{2}$", col): colclass="date"
        elif col.endswith("_info"): colclass="info"
        elif col.endswith("_remark"): colclass="remark"
        else: colclass=""
        th_attr=f' colname="{col}" colclass="{colclass}"'
        thead += f"<th{th_attr}>{html.escape(display_col)}</th>"
    thead += "</tr>"

    # 表体
    tbody = ""
    for _, row in df_excel.iterrows():
        tbody += "<tr>"
        for col in filtered_cols:
            val = row[col]
            if col=="rank": 
                tbody += f'<td class="rank">{int(val) if val!="" else ""}</td>'
                continue
            colclass="remark" if col.endswith("_remark") else "info" if col.endswith("_info") else "date" if re.match(r"\d{4}-\d{2}-\d{2}$",col) else ""
            td_attr=f' colname="{col}" colclass="{colclass}"'
            if colclass=="info":
                tbody += format_html_info(val)
                continue
            if colclass=="remark": 
                tbody += f'<td class="remark"{td_attr}>{html.escape(str(val))}</td>'
                continue
            if isinstance(val,str) and " / " not in val and not col.endswith("_info"):
                bk = html.escape(val.strip())
                if bk: tbody += f'<td class="bk"{td_attr} data-bk="{bk}" onclick="toggleBk(this)">{bk}</td>'
                else: tbody += f"<td{td_attr}></td>"
                continue
            tbody += f"<td{td_attr}>{html.escape(str(val))}</td>"
        tbody += "</tr>"

    table_html = f'<div class="table-wrap"><table><thead>{thead}</thead><tbody>{tbody}</tbody></table></div>'
    html_page = f"<!DOCTYPE html><html lang='zh-CN'><head><meta charset='utf-8'><title>Top</title>{STYLE}</head><body>{table_html}{SCRIPT}</body></html>"
    Path(OUT_HTML).write_text(html_page, encoding="utf-8")
    print(f"[OK] HTML 已生成：{OUT_HTML}")

    # ---------- AllData HTML 生成（显示全部 info，红绿显示涨跌幅） ----------
    df_all = pd.read_excel(OUTPUT_XLSX, sheet_name="AllData").fillna("")
    filtered_cols = [c for c in df_all.columns if not (c.endswith("_info") and not SHOW_INFO_COL) and not (c.endswith("_remark") and not SHOW_REMARK_COL)]
    display_cols = [shorten_col(c) for c in filtered_cols]

    thead = "<tr>"
    for col, display_col in zip(filtered_cols, display_cols):
        if col=="rank": colclass="rank"
        elif re.match(r"\d{4}-\d{2}-\d{2}$", col): colclass="date"
        elif col.endswith("_info"): colclass="info"
        elif col.endswith("_remark"): colclass="remark"
        else: colclass=""
        th_attr=f' colname="{col}" colclass="{colclass}"'
        thead += f"<th{th_attr}>{html.escape(display_col)}</th>"
    thead += "</tr>"

    tbody = ""
    for _, row in df_all.iterrows():
        tbody += "<tr>"
        for col in filtered_cols:
            val = row[col]
            colclass="remark" if col.endswith("_remark") else "info" if col.endswith("_info") else "date" if re.match(r"\d{4}-\d{2}-\d{2}$",col) else ""
            td_attr=f' colname="{col}" colclass="{colclass}"'

            if colclass=="info":
                # 获取涨跌幅（第一个数字）判断颜色
                if isinstance(val, str) and val.strip():
                    first_num = val.strip().split("/")[0].strip()
                    try:
                        pct = float(first_num)
                        color_class = "pos" if pct >=0 else "neg"
                    except:
                        color_class = ""
                else:
                    color_class = ""
                tbody += f'<td class="{color_class} info"{td_attr}>{html.escape(str(val))}</td>'
                continue

            if colclass=="remark": 
                tbody += f'<td class="{colclass}"{td_attr}>{html.escape(str(val))}</td>'
                continue
            if col=="rank":
                tbody += f'<td class="rank">{int(val) if val!="" else ""}</td>'
                continue
            if isinstance(val,str) and " / " not in val and not col.endswith("_info"):
                bk = html.escape(val.strip())
                if bk: tbody += f'<td class="bk"{td_attr} data-bk="{bk}" onclick="toggleBk(this)">{bk}</td>'
                else: tbody += f"<td{td_attr}></td>"
                continue
            tbody += f"<td{td_attr}>{html.escape(str(val))}</td>"
        tbody += "</tr>"

    table_html = f'<div class="table-wrap"><table><thead>{thead}</thead><tbody>{tbody}</tbody></table></div>'
    html_page = f"<!DOCTYPE html><html lang='zh-CN'><head><meta charset='utf-8'><title>Info</title>{STYLE}</head><body>{table_html}{SCRIPT}</body></html>"
    Path("html/bk_day_info.html").write_text(html_page, encoding="utf-8")
    print("[OK] AllData HTML 已生成：html/bk_day_info.html")


        # ---------- AllData Rank HTML（仅涨跌幅，完整排名） ----------
    if GEN_HTML_ALLDATA_RANK:
        df_rank = pd.read_excel(OUTPUT_XLSX, sheet_name="AllData").fillna("")

        # 仅保留 rank / 名称 / 日期列 / info
        filtered_cols = []
        for col in df_rank.columns:
            if col == "rank":
                filtered_cols.append(col)
            elif re.match(r"\d{4}-\d{2}-\d{2}$", col):
                filtered_cols.append(col)
            elif col.endswith("_info"):
                filtered_cols.append(col)

        display_cols = [shorten_col(c) for c in filtered_cols]

        # 表头
        thead = "<tr>"
        for col, display_col in zip(filtered_cols, display_cols):
            if col == "rank":
                colclass = "rank"
            elif re.match(r"\d{4}-\d{2}-\d{2}$", col):
                colclass = "date"
            elif col.endswith("_info"):
                colclass = "info"
            else:
                colclass = ""
            th_attr = f' colname="{col}" colclass="{colclass}"'
            thead += f"<th{th_attr}>{html.escape(display_col)}</th>"
        thead += "</tr>"

        # 表体
        tbody = ""
        for _, row in df_rank.iterrows():
            tbody += "<tr>"
            for col in filtered_cols:
                val = row[col]

                if col == "rank":
                    tbody += f'<td class="rank">{int(val) if val != "" else ""}</td>'
                    continue

                colclass = "info" if col.endswith("_info") else "date"
                td_attr = f' colname="{col}" colclass="{colclass}"'

                if colclass == "info":
                    tbody += format_html_info(val)
                    continue

                if isinstance(val, str) and " / " not in val:
                    bk = html.escape(val.strip())
                    if bk:
                        tbody += f'<td class="bk"{td_attr} data-bk="{bk}" onclick="toggleBk(this)">{bk}</td>'
                    else:
                        tbody += f"<td{td_attr}></td>"
                    continue

                tbody += f"<td{td_attr}>{html.escape(str(val))}</td>"

            tbody += "</tr>"

        table_html = f'<div class="table-wrap"><table><thead>{thead}</thead><tbody>{tbody}</tbody></table></div>'
        html_page = f"""<!DOCTYPE html>
    <html lang="zh-CN">
    <head>
    <meta charset="utf-8">
    <title>Rank</title>
    {STYLE}
    </head>
    <body>
    {table_html}
    {SCRIPT}
    </body>
    </html>
    """
        Path("html/bk_day_rank.html").write_text(html_page, encoding="utf-8")
        print("[OK] Rank HTML 已生成：html/bk_day_rank.html")


if __name__=="__main__":
    main()